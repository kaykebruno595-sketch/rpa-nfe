import streamlit as st
import xml.etree.ElementTree as ET
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="RPA - Gerador de Planilha de Nota", layout="centered")

st.title("📦 RPA Conversor de XML Copacker")
st.write("Atualize a planilha base e converta seus XMLs!")

# --- FUNÇÃO PARA LIMPEZA EXTREMA DE TEXTO ---
def limpar_texto_comparacao(texto):
    if pd.isna(texto):
        return ""
    txt = str(texto).upper().strip()
    txt = re.sub(r'[.\-\/\,\_\:\(\)]', '', txt)
    txt = txt.replace("LTDA", "").replace("S/A", "").replace("SA", "").replace("S.A", "")
    return " ".join(txt.split())

# --- 1. SEÇÃO DE UPLOAD DA PLANILHA BASE ---
st.header("1️⃣ Atualizar Planilha Base (Trânsito)")
arquivo_base_upload = st.file_uploader("Arraste aqui a planilha 'base_transito.xlsx' atualizada", type=["xlsx"])

df_base = None

if arquivo_base_upload is not None:
    try:
        df_base = pd.read_excel(arquivo_base_upload, sheet_name=0)
        df_base.columns = [str(col).strip().upper() for col in df_base.columns]
        
        col_nf = next((c for c in df_base.columns if "NF" in c), None)
        col_emit = next((c for c in df_base.columns if "EMIT" in c or "FORN" in c), None)
        col_mat = next((c for c in df_base.columns if "MAT" in c), None)
        col_desc_base = "XPROD" if "XPROD" in df_base.columns else next((c for c in df_base.columns if "DESC" in c or "PROD" in c), None)
        
        if col_nf and col_emit and col_mat:
            df_base[col_nf] = df_base[col_nf].apply(lambda x: str(int(float(x))) if re.match(r'^\d+\.\d+$', str(x)) else str(x).strip())
            
            st.success("✅ Planilha Base carregada e ativa na memória do site!")
            
            colunas_exibicao = [col_nf, col_emit, col_mat]
            if col_desc_base and col_desc_base not in colunas_exibicao:
                colunas_exibicao.append(col_desc_base)
                
            df_preview = df_base[colunas_exibicao].tail(5)
            st.dataframe(df_preview, use_container_width=True)
            
            st.session_state['cols_base'] = {'nf': col_nf, 'emit': col_emit, 'mat': col_mat}
        else:
            st.error("❌ Não foi possível encontrar as colunas necessárias (NF, EMIT/FORN, MAT) na planilha.")
            df_base = None

    except Exception as e:
        st.error(f"❌ Erro ao ler o arquivo Excel: {e}")
        df_base = None
else:
    st.info("💡 Aguardando o upload da planilha base para ativar a validação de códigos.")

st.write("---")

# --- 2. SEÇÃO DE UPLOAD DOS XMLS ---
st.header("2️⃣ Processar Arquivos XML")
arquivos_xml = st.file_uploader("Escolha os arquivos XML da nota", type=["xml"], accept_multiple_files=True)

if arquivos_xml:
    if df_base is None:
        st.error("❌ Erro impeditivo: Você precisa carregar uma Planilha Base válida no Passo 1 antes de processar os XMLs!")
    else:
        cols_b = st.session_state.get('cols_base')
        
        for arquivo in arquivos_xml:
            try:
                conteudo_xml = arquivo.read()
                raiz = ET.fromstring(conteudo_xml)
                
                ns = {'ns': 'http://www.portalfiscal.inf.br/nfe'}
                infNFe = raiz.find('.//ns:infNFe', ns)
                if infNFe is None:
                    infNFe = raiz.find('.//infNFe')
                    ns = {}
                
                if infNFe is None:
                    st.error(f"O arquivo {arquivo.name} não é uma NF-e válida.")
                    continue
                    
                ide = infNFe.find('ns:ide', ns) if ns else infNFe.find('ide')
                num_nota = ide.find('ns:nNF', ns).text if ns else ide.find('nNF').text
                num_nota_limpo = str(int(num_nota)) if num_nota.isdigit() else str(num_nota).strip()
                
                # --- COLETAR FORNECEDOR (XML) ---
                fornecedor_final = "Não Identificado"
                emit = infNFe.find('ns:emit', ns) if ns else infNFe.find('emit')
                if emit is not None:
                    xNome = emit.find('ns:xNome', ns) if ns else emit.find('xNome')
                    xFant = emit.find('ns:xFant', ns) if ns else emit.find('xFant')
                    if xNome is not None: fornecedor_final = xNome.text
                    elif xFant is not None: fornecedor_final = xFant.text

                # --- COLETAR CIDADE COM TRAVA (XML) ---
                cidade_final = "Outros / Não Encontrado"
                dest = infNFe.find('ns:dest', ns) if ns else infNFe.find('dest')
                if dest is not None:
                    enderDest = dest.find('ns:enderDest', ns) if ns else dest.find('enderDest')
                    if enderDest is not None:
                        xMun_node = enderDest.find('ns:xMun', ns) if ns else enderDest.find('xMun')
                        if xMun_node is not None:
                            cidade_xml_bruta = xMun_node.text.upper()
                            mapeamento_cidades = {
                                "CARIACICA": "Positive CO", "NATAL": "Natal", "POSITIVE": "Positive",
                                "SANTA LUZIA": "Santa Luzia", "ARAMA": "Arama", "LONDRINA": "Londrina", "DIADEMA": "Diadema"
                            }
                            cidade_final = next((v for k, v in mapeamento_cidades.items() if k in cidade_xml_bruta), cidade_xml_bruta.title())

                # --- COLETAR PESOS E VOLUMES (XML) ---
                transp = infNFe.find('ns:transp', ns) if ns else infNFe.find('transp')
                peso_liquido, peso_bruto, especie_volume, qtde_volume = 0.0, 0.0, "-", 0
                if transp is not None:
                    vol = transp.find('ns:vol', ns) if ns else transp.find('vol')
                    if vol is not None:
                        p_liq = vol.find('ns:pesoL', ns) if ns else vol.find('pesoL')
                        p_bru = vol.find('ns:pesoB', ns) if ns else vol.find('pesoB')
                        esp = vol.find('ns:esp', ns) if ns else vol.find('esp')
                        q_vol = vol.find('ns:qVol', ns) if ns else vol.find('qVol')
                        if p_liq is not None: peso_liquido = float(p_liq.text)
                        if p_bru is not None: peso_bruto = float(p_bru.text)
                        if esp is not None: especie_volume = esp.text
                        if q_vol is not None: qtde_volume = int(q_vol.text)

                # --- FILTRAR LINHAS DA NOTA NA BASE ---
                df_nota_especifica = df_base[df_base[cols_b['nf']] == num_nota_limpo]
                linhas_nota_base = []
                
                if not df_nota_especifica.empty:
                    fornecedor_xml_ultra_limpo = limpar_texto_comparacao(fornecedor_final)
                    for _, linha_base in df_nota_especifica.iterrows():
                        fornecedor_base_ultra_limpo = limpar_texto_comparacao(linha_base[cols_b['emit']])
                        if fornecedor_base_ultra_limpo in fornecedor_xml_ultra_limpo or fornecedor_xml_ultra_limpo in fornecedor_base_ultra_limpo:
                            linhas_nota_base.append(str(linha_base[cols_b['mat']]).strip())

                # --- MONTAGEM DOS ITENS ---
                lista_produtos = []
                itens_xml = infNFe.findall('ns:det', ns) if ns else infNFe.findall('det')

                for idx, item in enumerate(itens_xml):
                    prod = item.find('ns:prod', ns) if ns else item.find('prod')
                    
                    if prod is not None:
                        codigo_original_xml = prod.find('ns:cProd', ns).text if ns else prod.find('cProd').text
                        nome_produto = prod.find('ns:xProd', ns).text if ns else prod.find('xProd').text
                        umb = prod.find('ns:uCom', ns).text if ns else prod.find('uCom').text  
                        quantidade = float(prod.find('ns:qCom', ns).text if ns else prod.find('qCom').text)
                        valor_unitario = float(prod.find('ns:vUnCom', ns).text if ns else prod.find('vUnCom').text)
                        valor_total_item = float(prod.find('ns:vProd', ns).text if ns else prod.find('vProd').text)
                        
                        codigo_substituto = None
                        if idx < len(linhas_nota_base):
                            codigo_substituto = linhas_nota_base[idx]
                        
                        codigo_final_item = codigo_substituto if (codigo_substituto and codigo_substituto.lower() != 'nan') else codigo_original_xml
                        
                        imposto = item.find('ns:imposto', ns) if ns else item.find('imposto')
                        valor_icms_num, valor_ipi_num = 0.0, 0.0
                        
                        if imposto is not None:
                            icms_bloco = imposto.find('.//ns:ICMS', ns) if ns else imposto.find('.//ICMS')
                            if icms_bloco is not None:
                                for sub_tag in icms_bloco.iter():
                                    if sub_tag.tag.endswith('pICMS'):
                                        valor_icms_num = float(sub_tag.text)
                                        break
                                        
                            ipi_bloco = imposto.find('.//ns:IPI', ns) if ns else imposto.find('.//IPI')
                            if ipi_bloco is not None:
                                for sub_tag in ipi_bloco.iter():
                                    if sub_tag.tag.endswith('pIPI'):
                                        valor_ipi_num = float(sub_tag.text)
                                        break
                        
                        lista_produtos.append({
                            "CODIGO": codigo_final_item,
                            "DESCRIÇÃO": nome_produto,
                            "NOTA FISCAL": num_nota_limpo,
                            "UMB": umb,
                            "QTDE": quantidade,
                            "VLR. UNT.": valor_unitario,
                            "VLR. TT.": valor_total_item,
                            "ICMS": valor_icms_num,
                            "IPI": valor_ipi_num
                        })

                # --- NOVO DESIGN DA PLANILHA (OPENPYXL) ---
                wb = Workbook()
                ws = wb.active
                ws.title = f"NF {num_nota_limpo}"
                ws.views.sheetView[0].showGridLines = True
                
                # Cores de Estilo do cabeçalho
                cor_azul_escuro, col_azul_claro = "1B365D", "F0F4F8"
                fill_header = PatternFill(start_color=cor_azul_escuro, end_color=cor_azul_escuro, fill_type="solid")
                fill_sub_header = PatternFill(start_color=col_azul_claro, end_color=col_azul_claro, fill_type="solid")
                font_branca_negrito = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                font_preta_negrito = Font(name="Calibri", size=11, bold=True, color="000000")
                font_normal = Font(name="Calibri", size=11)
                
                border_fina = Border(
                    left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
                )
                
                # Linha 1: Título Geral do Bloco
                ws.merge_cells("A1:N1")
                ws["A1"] = "DADOS MATERIAIS"
                ws["A1"].fill = fill_header
                ws["A1"].font = font_branca_negrito
                ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 25
                
                # Linha 2: Duplo cabeçalho para "LANÇAR NF" e "INDUSTRIALIZAÇÃO"
                ws.merge_cells("L2:M2")
                ws["L2"] = "LANÇAR NF"
                ws["L2"].fill = fill_header
                ws["L2"].font = font_branca_negrito
                ws["L2"].alignment = Alignment(horizontal="center", vertical="center")
                
                ws.cell(row=2, column=14, value="INDUSTRIALIZAÇÃO").fill = fill_header
                ws.cell(row=2, column=14).font = font_branca_negrito
                ws.cell(row=2, column=14).alignment = Alignment(horizontal="center", vertical="center")
                
                # Linha 3: Cabeçalhos Reais das Colunas Exatas da Imagem
                colunas = [
                    "CÓDIGO", "DESCRIÇÃO", "CENTRO", "NOTA FISCAL", "DEPÓSITO", 
                    "UMB", "QTDE", "VLR. UNT.", "VLR. TT.", "ICMS", "IPI", 
                    "OC/ITEM", "OC/ITEM", "OC/ITEM"
                ]
                
                for col_idx, texto_coluna in enumerate(colunas, 1):
                    celula = ws.cell(row=3, column=col_idx, value=texto_coluna)
                    celula.fill = fill_header
                    celula.font = font_branca_negrito
                    celula.alignment = Alignment(horizontal="center", vertical="center")
                    celula.border = border_fina
                ws.row_dimensions[3].height = 22
                
                # Preenchimento dos dados dinâmicos e das colunas vazias solicitadas
                linha_atual = 4
                for prod in lista_produtos:
                    ws.cell(row=linha_atual, column=1, value=prod["CODIGO"]).alignment = Alignment(horizontal="center")
                    ws.cell(row=linha_atual, column=2, value=prod["DESCRIÇÃO"]).alignment = Alignment(horizontal="left")
                    
                    # Coluna 3: CENTRO (Vazia)
                    ws.cell(row=linha_atual, column=3, value="") 
                    
                    try:
                        ws.cell(row=linha_atual, column=4, value=int(prod["NOTA FISCAL"])).alignment = Alignment(horizontal="center")
                    except ValueError:
                        ws.cell(row=linha_atual, column=4, value=prod["NOTA FISCAL"]).alignment = Alignment(horizontal="center")
                        
                    # Coluna 5: DEPÓSITO (Vazia)
                    ws.cell(row=linha_atual, column=5, value="")
                    
                    ws.cell(row=linha_atual, column=6, value=prod["UMB"]).alignment = Alignment(horizontal="center")
                    ws.cell(row=linha_atual, column=7, value=prod["QTDE"]).number_format = '#,##0.00'
                    ws.cell(row=linha_atual, column=8, value=prod["VLR. UNT."]).number_format = '#,##0.00'
                    
                    celula_vlr_tt = ws.cell(row=linha_atual, column=9, value=prod["VLR. TT."])
                    celula_vlr_tt.number_format = '"R$ " #,##0.00;"R$ " -'
                    
                    celula_icms = ws.cell(row=linha_atual, column=10, value=prod["ICMS"])
                    celula_icms.number_format = '#,##0.00'
                    celula_icms.alignment = Alignment(horizontal="center")
                    
                    celula_ipi = ws.cell(row=linha_atual, column=11, value=prod["IPI"])
                    celula_ipi.number_format = '#,##0.00'
                    celula_ipi.alignment = Alignment(horizontal="center")
                    
                    # Colunas 12, 13 e 14: Campos OC/ITEM (Vazios)
                    ws.cell(row=linha_atual, column=12, value="")
                    ws.cell(row=linha_atual, column=13, value="")
                    ws.cell(row=linha_atual, column=14, value="")
                    
                    for c in range(1, 15):
                        ws.cell(row=linha_atual, column=c).font = font_normal
                        ws.cell(row=linha_atual, column=c).border = border_fina
                    ws.row_dimensions[linha_atual].height = 20
                    linha_atual += 1
                    
                # Espaçamento e estruturação dos Blocos Inferiores (Dados Fiscais)
                linha_atual += 2
                ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=3)
                celula_fiscal_titulo = ws.cell(row=linha_atual, column=1, value="DADOS FISCAIS")
                celula_fiscal_titulo.fill = fill_header
                celula_fiscal_titulo.font = font_branca_negrito
                ws.row_dimensions[linha_atual].height = 22
                linha_atual += 1
                
                dados_fiscais_valores = [
                    ("PESO BRUTO", peso_bruto),
                    ("PESO LÍQUIDO", peso_liquido),
                    ("ESPÉCIE DE VOLUME", especie_volume),
                    ("QTDE VOLUME", qtde_volume)
                ]
                
                for label, valor in dados_fiscais_valores:
                    ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
                    c_label = ws.cell(row=linha_atual, column=1, value=label)
                    c_label.fill = fill_sub_header
                    c_label.font = font_preta_negrito
                    
                    c_valor = ws.cell(row=linha_atual, column=3, value=valor)
                    if isinstance(valor, float): c_valor.number_format = '#,##0.00'
                    c_valor.font = font_normal
                    c_valor.border = border_fina
                    c_valor.alignment = Alignment(horizontal="left")
                    
                    for col_borda in range(1, 4):
                        ws.cell(row=linha_atual, column=col_borda).border = border_fina
                    ws.row_dimensions[linha_atual].height = 18
                    linha_atual += 1
                
                # Estruturação dos Campos de Transporte Inferiores (Vazios como no modelo)
                linha_atual += 1
                ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=3)
                celula_transp_titulo = ws.cell(row=linha_atual, column=1, value="DADOS DE TRANSPORTE")
                celula_transp_titulo.fill = fill_header
                celula_transp_titulo.font = font_branca_negrito
                ws.row_dimensions[linha_atual].height = 22
                linha_atual += 1
                
                dados_transporte_labels = ["FRETE PAGO", "CÓDIGO", "TRANSPORTADORA"]
                for label in dados_transporte_labels:
                    ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=2)
                    c_label = ws.cell(row=linha_atual, column=1, value=label)
                    c_label.fill = fill_sub_header
                    c_label.font = font_preta_negrito
                    
                    c_vazio = ws.cell(row=linha_atual, column=3, value="")
                    for col_borda in range(1, 4):
                        ws.cell(row=linha_atual, column=col_borda).border = border_fina
                    ws.row_dimensions[linha_atual].height = 18
                    linha_atual += 1
                
                # Auto-ajuste inteligente de tamanho das colunas
                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.row in [1, 2]: continue  
                        if cell.value: max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
                ws.column_dimensions['A'].width = 16
                ws.column_dimensions['B'].width = 45
                ws.column_dimensions['D'].width = 16
                
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                fornecedor_limpo = re.sub(r'[\\/*?:"<>|]', "", fornecedor_final).strip()
                
                st.download_button(
                    label=f"📥 Baixar nota {num_nota_limpo} - {fornecedor_limpo}",
                    data=buffer,
                    file_name=f"PLANILHA_NOTA_{num_nota_limpo}_{fornecedor_limpo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            except Exception as e:
                st.error(f"Erro ao processar o arquivo {arquivo.name}: {e}")
